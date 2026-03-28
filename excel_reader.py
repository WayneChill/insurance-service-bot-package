"""
excel_reader.py ── 從 Google Drive「保服資料」資料夾讀取 42003.xlsx / 42004.xlsx
並提供每日早報所需的壽險壽星 / 保單周年統計
"""
import os
import io
import json
import base64
import re
from datetime import datetime
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
DRIVE_FOLDER = "保服資料"
LIFE_FILE     = "42003.xlsx"
PROPERTY_FILE = "42004.xlsx"

_POLICY_SUFFIX = re.compile(r"\d{7}")


def get_creds():
    b64 = os.environ.get("GOOGLE_CREDENTIALS_B64", "")
    if b64:
        info = json.loads(base64.b64decode(b64).decode("utf-8"))
    else:
        path = os.environ.get("GOOGLE_CREDENTIALS_FILE", "credentials.json")
        with open(path, encoding="utf-8") as f:
            info = json.load(f)
    return Credentials.from_service_account_info(info, scopes=SCOPES)


def _find_folder_id(drive):
    folder_id = os.environ.get("DRIVE_FOLDER_ID", "")
    if folder_id:
        return folder_id
    res = drive.files().list(
        q=f"name='{DRIVE_FOLDER}' and mimeType='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id)"
    ).execute()
    folders = res.get("files", [])
    if not folders:
        raise RuntimeError(f"找不到 Google Drive 資料夾「{DRIVE_FOLDER}」")
    return folders[0]["id"]


def download_excel(filename: str) -> io.BytesIO:
    creds = get_creds()
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)
    folder_id = _find_folder_id(drive)
    file_res = drive.files().list(
        q=f"name='{filename}' and '{folder_id}' in parents and trashed=false",
        fields="files(id)",
        pageSize=1
    ).execute()
    files = file_res.get("files", [])
    if not files:
        raise RuntimeError(f"找不到 {filename}")
    req = drive.files().get_media(fileId=files[0]["id"])
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf


# ── 通用工具 ──────────────────────────────────────────────
def safe_str(val, default=""):
    if val is None:
        return default
    if isinstance(val, float):
        return str(int(val))
    if isinstance(val, int):
        return str(val)
    return str(val).strip().replace("\n", "")


def safe_get(row, idx, default=""):
    try:
        return safe_str(row[idx], default)
    except (IndexError, TypeError):
        return default


def _clean_policy(policy, start_date):
    if not policy or not start_date or len(start_date) < 3:
        return policy
    year = start_date[:3]
    for y in [year, str(int(year) + 1)]:
        idx = policy.find(y)
        if idx > 0:
            return policy[:idx]
    return policy


def roc_to_ad(val):
    """民國日期（7碼整數）→ datetime.date，失敗回傳 None"""
    try:
        s = str(int(val)).zfill(7)
        return datetime(int(s[0:3]) + 1911, int(s[3:5]), int(s[5:7])).date()
    except Exception:
        return None


# ── 保服查詢：查客戶資料 ───────────────────────────────────
def parse_life_excel(buf: io.BytesIO, name: str) -> list:
    wb = load_workbook(buf, read_only=True)
    ws = wb.active
    results = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 8:
            continue
        if not row or len(row) < 25:
            continue
        policy_raw = safe_get(row, 0)
        if not policy_raw or policy_raw.startswith("附約") or policy_raw == "保單號碼":
            continue
        insured   = safe_get(row, 21)
        applicant = safe_get(row, 19)
        if name not in [insured, applicant]:
            continue
        key        = insured if insured else applicant
        policy_num = _clean_policy(policy_raw, safe_get(row, 10))
        if not policy_num:
            continue
        if key not in results:
            idno_raw = safe_get(row, 23)
            results[key] = {
                "name": key,
                "applicant": applicant if applicant != key else "",
                "idno": idno_raw[:10] if len(idno_raw) >= 10 else idno_raw,
                "tel": safe_get(row, 34),
                "addr": safe_get(row, 36),
                "policies": [],
            }
        existing = {p["policy_num"] for p in results[key]["policies"]}
        if policy_num not in existing:
            results[key]["policies"].append({
                "type": "壽險",
                "company": safe_get(row, 2),
                "policy_num": policy_num,
                "product": safe_get(row, 7),
                "status": safe_get(row, 32),
            })
    return list(results.values())


def parse_property_excel(buf: io.BytesIO, name: str) -> list:
    wb = load_workbook(buf, read_only=True)
    ws = wb.active
    results = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5:
            continue
        if not row or len(row) < 20:
            continue
        policy_raw = safe_get(row, 1)
        if not policy_raw or len(policy_raw) < 5:
            continue
        company = safe_get(row, 3)
        if not company or company == "險種代號":
            continue
        insured   = safe_get(row, 10)
        applicant = safe_get(row, 9)
        if name not in [insured, applicant]:
            continue
        key        = insured if insured else applicant
        policy_num = _clean_policy(policy_raw, safe_get(row, 18))
        if not policy_num:
            continue
        if key not in results:
            results[key] = {
                "name": key,
                "applicant": applicant if applicant != key else "",
                "idno": safe_get(row, 11)[:10],
                "tel": safe_get(row, 23),
                "addr": safe_get(row, 25),
                "policies": [],
            }
        existing = {p["policy_num"] for p in results[key]["policies"]}
        if policy_num not in existing:
            results[key]["policies"].append({
                "type": "產險",
                "company": company,
                "policy_num": policy_num,
                "product": safe_get(row, 16),
                "status": safe_get(row, 21),
            })
    return list(results.values())


def search_client(name: str) -> list:
    """查詢客戶（合併壽險 + 產險）"""
    combined = {}

    try:
        buf = download_excel(LIFE_FILE)
        for r in parse_life_excel(buf, name):
            k = r["name"]
            if k not in combined:
                combined[k] = r
            else:
                existing = {p["policy_num"] for p in combined[k]["policies"]}
                for p in r["policies"]:
                    if p["policy_num"] not in existing:
                        combined[k]["policies"].append(p)
        print("Life OK", flush=True)
    except Exception as e:
        print(f"[WARN] 壽險: {e}", flush=True)

    try:
        buf = download_excel(PROPERTY_FILE)
        for r in parse_property_excel(buf, name):
            k = r["name"]
            if k not in combined:
                combined[k] = r
            else:
                existing = {p["policy_num"] for p in combined[k]["policies"]}
                for p in r["policies"]:
                    if p["policy_num"] not in existing:
                        combined[k]["policies"].append(p)
        print("Property OK", flush=True)
    except Exception as e:
        print(f"[WARN] 產險: {e}", flush=True)

    return list(combined.values())


# ── 每日早報：壽險壽星 / 保單周年 ─────────────────────────
def get_life_daily_stats() -> dict:
    """
    回傳 {"birthday_count": N, "anniversary_count": N}
    壽星 = 出生月日 == 今日月日（欄位索引依 42003.xlsx 實際格式調整）
    保單周年 = 生效月日 == 今日月日
    """
    stats = {"birthday_count": 0, "anniversary_count": 0}
    try:
        buf   = download_excel(LIFE_FILE)
        wb    = load_workbook(buf, read_only=True)
        ws    = wb.active
        today = datetime.today()
        seen_insured = set()
        seen_policy  = set()

        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i < 8:
                continue
            if not row or len(row) < 35:
                continue
            policy_raw = safe_get(row, 0)
            if not policy_raw or policy_raw.startswith("附約") or policy_raw == "保單號碼":
                continue

            # 壽星：欄位 22 = 被保人出生日（民國7碼）
            insured = safe_get(row, 21)
            dob_val = safe_get(row, 22)
            if insured and insured not in seen_insured and dob_val:
                dob = roc_to_ad(dob_val)
                if dob and dob.month == today.month and dob.day == today.day:
                    stats["birthday_count"] += 1
                    seen_insured.add(insured)

            # 保單周年：欄位 10 = 保險起日（民國7碼）
            policy_num = _clean_policy(policy_raw, safe_get(row, 10))
            if policy_num and policy_num not in seen_policy:
                start_val = safe_get(row, 10)
                start_date = roc_to_ad(start_val)
                if start_date and start_date.month == today.month and start_date.day == today.day:
                    stats["anniversary_count"] += 1
                    seen_policy.add(policy_num)

    except Exception as e:
        print(f"[WARN] 壽險早報統計失敗: {e}")
    return stats


# ── 每日早報：產險各群件數 ────────────────────────────────
def get_property_daily_stats(statuses: dict) -> dict:
    """
    從 42004.xlsx 計算急件/追蹤/新件，再加上 Sheets 內的延後件數
    statuses: get_property_status() 回傳的 dict
    """
    stats = {"urgent": 0, "track": 0, "new": 0, "delay": 0}
    try:
        import pandas as pd
        buf = download_excel(PROPERTY_FILE)
        df  = pd.read_excel(buf, header=3)
        df.columns = df.columns.str.strip()
        df = df[df["保單號碼"].notna()]
        df = df[~df["保單號碼"].astype(str).str.contains("險種代號|附約")]

        def _roc(v):
            try:
                s = str(int(v)).zfill(7)
                return pd.Timestamp(int(s[0:3]) + 1911, int(s[3:5]), int(s[5:7]))
            except Exception:
                return pd.NaT

        df["到期日"]   = df["保險迄日"].apply(_roc)
        today          = pd.Timestamp(datetime.today().date())
        df["剩餘天數"] = (df["到期日"] - today).dt.days
        active = df[
            df["剩餘天數"].notna() &
            df["剩餘天數"].between(0, 60) &
            df["狀態"].astype(str).str.contains("正常")
        ]
        skip = {"續保完成", "不續保"}
        delay_statuses = {"延後3天", "延後7天"}
        for _, row in active.iterrows():
            pid    = str(row["保單號碼"]).strip()
            cur    = statuses.get(pid, {}).get("status", "")
            if cur in skip:
                continue
            if cur in delay_statuses:
                stats["delay"] += 1
                continue
            days = int(row["剩餘天數"])
            if days <= 10:
                stats["urgent"] += 1
            elif days <= 30:
                stats["track"] += 1
            else:
                stats["new"] += 1
    except Exception as e:
        print(f"[WARN] 產險早報統計失敗: {e}")
    return stats
